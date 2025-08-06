"""
stmate.xlsx 파일 구조 분석 (openpyxl 직접 사용)
"""
import sys
import io

# UTF-8 인코딩 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_stmate_with_openpyxl():
    """openpyxl로 직접 stmate.xlsx 분석"""
    file_path = 'stmate.xlsx'
    
    try:
        import openpyxl
        
        # data_only=True로 수식 대신 값만 읽기
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        print(f"시트 목록: {sheets}")
        
        for sheet_name in sheets:
            print(f"\n{'='*50}")
            print(f"시트: {sheet_name}")
            print('='*50)
            
            try:
                ws = wb[sheet_name]
                print(f"크기: {ws.max_row} x {ws.max_column}")
                
                # 처음 20행, 10열만 읽기
                print("\n처음 20행 내용:")
                for row_idx in range(1, min(21, ws.max_row + 1)):
                    row_content = []
                    for col_idx in range(1, min(11, ws.max_column + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            content = str(cell.value).strip()
                            if content:
                                row_content.append(f"[{col_idx}]:{content[:30]}")
                    if row_content:
                        print(f"행{row_idx}: {' | '.join(row_content)}")
                
                # 호표 패턴 찾기
                print(f"\n호표 패턴 검색:")
                import re
                patterns = [
                    r'제\d+호표',      # 제N호표
                    r'호표\s*\d+',     # 호표 N
                    r'\d+\.\s*',       # N. 형식
                    r'산근\s*\d+',     # 산근 N
                    r'#\d+',           # #N
                    r'\(\s*산근\s*\d+\s*\)'  # (산근 N)
                ]
                
                found_patterns = []
                
                for row_idx in range(1, min(101, ws.max_row + 1)):  # 처음 100행만 검색
                    for col_idx in range(1, min(11, ws.max_column + 1)):  # 처음 10열만
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell_str = str(cell.value).strip()
                            for pattern in patterns:
                                if re.search(pattern, cell_str):
                                    found_patterns.append({
                                        'row': row_idx,
                                        'col': col_idx,
                                        'pattern': pattern,
                                        'text': cell_str[:50]
                                    })
                
                if found_patterns:
                    print("발견된 호표 패턴:")
                    for p in found_patterns[:10]:  # 처음 10개만
                        print(f"  행{p['row']}, 열{p['col']}: {p['pattern']} -> {p['text']}")
                else:
                    print("호표 패턴을 찾을 수 없습니다.")
                
                # 단가산출 관련 텍스트 찾기
                print(f"\n'단가산출' 관련 텍스트 검색:")
                for row_idx in range(1, min(51, ws.max_row + 1)):  # 처음 50행만
                    for col_idx in range(1, min(11, ws.max_column + 1)):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell_str = str(cell.value).strip()
                            if '단가산출' in cell_str or '산출' in cell_str or '내역' in cell_str:
                                print(f"  행{row_idx}, 열{col_idx}: {cell_str[:50]}")
                
            except Exception as e:
                print(f"시트 '{sheet_name}' 분석 오류: {e}")
    
    except Exception as e:
        print(f"파일 분석 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    analyze_stmate_with_openpyxl()