"""
단가산출 통합 파서
모든 파일의 단가산출 데이터를 추출하여 통일된 JSON 구조로 변환
"""
import pandas as pd
import json
import os
import re
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

class PriceCalculationParser:
    """단가산출 통합 파서"""
    
    def __init__(self):
        os.makedirs('result', exist_ok=True)
    
    def parse_test1_price(self):
        """test1.xlsx 단가산출 파싱"""
        print("\n" + "="*70)
        print("test1.xlsx 단가산출 파싱")
        print("="*70)
        
        df = pd.read_excel('test1.xlsx', sheet_name='단가산출_산근', header=None)
        print(f"시트 크기: {df.shape}")
        
        # 단가산출 항목 찾기 (숫자.제목 패턴)
        price_items = []
        
        for i in range(len(df)):
            cell = df.iloc[i, 0]
            if pd.notna(cell):
                cell_str = str(cell).strip()
                
                # 숫자.제목 패턴 찾기 (예: "1.아스팔트포장깨기")
                match = re.match(r'^(\d+)\.(.+)', cell_str)
                if match:
                    item_no = match.group(1)
                    item_name = match.group(2).strip()
                    
                    # 단위 정보 추출 [㎥], [㎡] 등
                    unit_match = re.search(r'\[([^\]]+)\]', item_name)
                    unit = unit_match.group(1) if unit_match else ''
                    
                    # 단위 제거한 품명
                    clean_name = re.sub(r'\s*\[[^\]]+\]', '', item_name).strip()
                    
                    price_items.append({
                        'row': i,
                        'no': item_no,
                        'name': clean_name,
                        'unit': unit,
                        'full_name': item_name
                    })
                    print(f"  발견: {item_no}. {clean_name} [{unit}]")
        
        print(f"단가산출 항목 발견: {len(price_items)}개")
        
        # JSON 결과 생성
        result = self._create_price_result('test1.xlsx', '단가산출_산근', price_items, df, 'test1')
        
        # 저장
        self._save_result(result, 'result/test1_price.json')
        
        return result
    
    def parse_est_price(self):
        """est.xlsx 단가산출 파싱"""
        print("\n" + "="*70)
        print("est.xlsx 단가산출 파싱")
        print("="*70)
        
        df = pd.read_excel('est.xlsx', sheet_name='단가산출', header=None)
        print(f"시트 크기: {df.shape}")
        
        # 단가산출 항목 찾기
        price_items = []
        
        for i in range(len(df)):
            for j in range(min(2, len(df.columns))):
                cell = df.iloc[i, j]
                if pd.notna(cell):
                    cell_str = str(cell).strip()
                    
                    # 숫자.제목 패턴 찾기
                    match = re.match(r'^(\d+)\.(.+)', cell_str)
                    if match:
                        item_no = match.group(1)
                        item_name = match.group(2).strip()
                        
                        # 합계, 재료비, 노무비, 경비 추출 (같은 행의 다른 컬럼들)
                        total = ''
                        labor = ''
                        material = ''
                        expense = ''
                        
                        if len(df.columns) > 2:
                            total_cell = df.iloc[i, 2]  # 합계
                            if pd.notna(total_cell):
                                total = str(total_cell).strip()
                        
                        if len(df.columns) > 3:
                            labor_cell = df.iloc[i, 3]  # 노무비
                            if pd.notna(labor_cell):
                                labor = str(labor_cell).strip()
                        
                        if len(df.columns) > 4:
                            material_cell = df.iloc[i, 4]  # 재료비
                            if pd.notna(material_cell):
                                material = str(material_cell).strip()
                        
                        if len(df.columns) > 5:
                            expense_cell = df.iloc[i, 5]  # 경비
                            if pd.notna(expense_cell):
                                expense = str(expense_cell).strip()
                        
                        price_items.append({
                            'row': i,
                            'no': item_no,
                            'name': item_name,
                            'total': total,
                            'labor': labor,
                            'material': material,
                            'expense': expense
                        })
                        print(f"  발견: {item_no}. {item_name[:40]}...")
                        break
        
        print(f"단가산출 항목 발견: {len(price_items)}개")
        
        # JSON 결과 생성
        result = self._create_price_result('est.xlsx', '단가산출', price_items, df, 'est')
        
        # 저장
        self._save_result(result, 'result/est_price.json')
        
        return result
    
    def parse_sgs_price(self):
        """sgs.xls 단가산출 파싱"""
        print("\n" + "="*70)
        print("sgs.xls 단가산출 파싱")
        print("="*70)
        
        df = pd.read_excel('sgs.xls', sheet_name='단가산출', header=None, engine='xlrd')
        print(f"시트 크기: {df.shape}")
        
        # 단가산출 항목 찾기 (산근 N 호표 패턴)
        price_items = []
        
        for i in range(len(df)):
            cell = df.iloc[i, 0]
            if pd.notna(cell):
                cell_str = str(cell).strip()
                
                # "산근 N 호표" 패턴 찾기
                match = re.search(r'산근\s*(\d+)\s*호표\s*[：:]\s*(.+)', cell_str)
                if match:
                    item_no = match.group(1)
                    item_name = match.group(2).strip()
                    
                    price_items.append({
                        'row': i,
                        'no': item_no,
                        'name': item_name,
                        'full_text': cell_str
                    })
                    print(f"  발견: 산근{item_no} - {item_name[:40]}...")
        
        print(f"단가산출 항목 발견: {len(price_items)}개")
        
        # JSON 결과 생성
        result = self._create_price_result('sgs.xls', '단가산출', price_items, df, 'sgs')
        
        # 저장
        self._save_result(result, 'result/sgs_price.json')
        
        return result
    
    def parse_construction1_price(self):
        """건축구조내역.xlsx 중기단가산출서 파싱"""
        print("\n" + "="*70)
        print("건축구조내역.xlsx 중기단가산출서 파싱")
        print("="*70)
        
        df = pd.read_excel('건축구조내역.xlsx', sheet_name='중기단가산출서', header=None)
        print(f"시트 크기: {df.shape}")
        
        # 단가산출 항목 찾기 (헤더 행 3 이후, 품목명이 있는 행)
        price_items = []
        item_no = 1
        
        for i in range(3, len(df)):  # 헤더 이후부터
            cell = df.iloc[i, 0]  # 첫 번째 컬럼 (산출내역)
            if pd.notna(cell):
                cell_str = str(cell).strip()
                
                # 빈 행이거나 소계, 합계 등이 아닌 품목명인 경우
                if cell_str and not cell_str.startswith(' ') and '소  계' not in cell_str and '총  계' not in cell_str and 'Q ' not in cell_str and 'HR ' not in cell_str:
                    
                    # 해당 행의 재료비, 노무비, 경비, 합계 추출
                    material = ''
                    labor = ''
                    expense = ''
                    total = ''
                    
                    if len(df.columns) > 1:
                        material_cell = df.iloc[i, 1]
                        if pd.notna(material_cell):
                            material = str(material_cell).strip()
                    
                    if len(df.columns) > 2:
                        labor_cell = df.iloc[i, 2]
                        if pd.notna(labor_cell):
                            labor = str(labor_cell).strip()
                    
                    if len(df.columns) > 3:
                        expense_cell = df.iloc[i, 3]
                        if pd.notna(expense_cell):
                            expense = str(expense_cell).strip()
                    
                    if len(df.columns) > 4:
                        total_cell = df.iloc[i, 4]
                        if pd.notna(total_cell):
                            total = str(total_cell).strip()
                    
                    # 유효한 데이터가 있는 경우만 추가
                    if any([material, labor, expense, total]) and any(val != '0' for val in [material, labor, expense, total]):
                        price_items.append({
                            'row': i,
                            'no': str(item_no),
                            'name': cell_str,
                            'material': material,
                            'labor': labor,
                            'expense': expense,
                            'total': total
                        })
                        print(f"  발견: {item_no}. {cell_str[:50]}...")
                        item_no += 1
        
        print(f"단가산출 항목 발견: {len(price_items)}개")
        
        # JSON 결과 생성
        result = self._create_price_result('건축구조내역.xlsx', '중기단가산출서', price_items, df, 'construction1')
        
        # 저장
        self._save_result(result, 'result/construction1_price.json')
        
        return result
    
    def parse_stmate_price(self):
        """stmate.xlsx 일위대가_산근 단가산출 파싱"""
        print("\n" + "="*70)
        print("stmate.xlsx 일위대가_산근 단가산출 파싱")
        print("="*70)
        
        # repaired 파일 사용
        wb = load_workbook('stmate_repaired.xlsx', read_only=True, data_only=True)
        ws = wb['일위대가_산근']
        
        # DataFrame으로 변환
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        df = pd.DataFrame(data)
        
        print(f"시트 크기: {df.shape}")
        
        # 단가산출 항목 찾기 (#N 패턴)
        price_items = []
        
        for i in range(len(df)):
            cell = df.iloc[i, 0]
            if pd.notna(cell):
                cell_str = str(cell).strip()
                
                # #N 패턴 찾기 (예: "#1 폐기물 상차 | 굴삭기 1.0㎥|㎥")
                match = re.match(r'^#(\d+)\s+(.+?)(?:\s*\|\s*([^|]+)\s*\|\s*([^|]+))?$', cell_str)
                if match:
                    item_no = match.group(1)
                    item_name = match.group(2).strip()
                    spec = match.group(3).strip() if match.group(3) else ''
                    unit = match.group(4).strip() if match.group(4) else ''
                    
                    # 같은 행의 합계, 재료비, 노무비, 경비 추출
                    total = str(df.iloc[i, 1]).strip() if len(df.columns) > 1 and pd.notna(df.iloc[i, 1]) else ''
                    material = str(df.iloc[i, 2]).strip() if len(df.columns) > 2 and pd.notna(df.iloc[i, 2]) else ''
                    labor = str(df.iloc[i, 3]).strip() if len(df.columns) > 3 and pd.notna(df.iloc[i, 3]) else ''
                    expense = str(df.iloc[i, 4]).strip() if len(df.columns) > 4 and pd.notna(df.iloc[i, 4]) else ''
                    
                    price_items.append({
                        'row': i,
                        'no': item_no,
                        'name': item_name,
                        'spec': spec,
                        'unit': unit,
                        'total': total,
                        'material': material,
                        'labor': labor,
                        'expense': expense
                    })
                    print(f"  발견: #{item_no} {item_name} | {spec} | {unit}")
        
        wb.close()
        print(f"단가산출 항목 발견: {len(price_items)}개")
        
        # JSON 결과 생성
        result = self._create_price_result('stmate.xlsx', '일위대가_산근', price_items, df, 'stmate')
        
        # 저장
        self._save_result(result, 'result/stmate_price.json')
        
        return result
    
    def _create_price_result(self, file_name: str, sheet_name: str, price_items: List, 
                           df: pd.DataFrame, file_type: str) -> Dict:
        """단가산출 JSON 결과 생성"""
        result = {
            'file': file_name,
            'sheet': sheet_name,
            'type': '단가산출',
            'total_price_items': len(price_items),
            'price_data': []
        }
        
        for item in price_items:
            # 상세 계산 과정 추출 (해당 항목 이후의 행들)
            start_row = item['row']
            end_row = start_row + 50  # 최대 50행까지 확인
            
            calculation_details = []
            
            # 다음 항목 시작점 찾기
            for next_item in price_items:
                if next_item['row'] > start_row:
                    end_row = min(end_row, next_item['row'])
                    break
            
            # 계산 과정 추출
            for row_idx in range(start_row + 1, min(end_row, len(df))):
                detail_row = self._extract_calculation_detail(df, row_idx, file_type)
                if detail_row:
                    calculation_details.append(detail_row)
            
            # 기본 구조
            price_item = {
                'item_no': item['no'],
                'item_name': item['name'],
                'position': {
                    'start_row': start_row,
                    'end_row': end_row,
                    'total_rows': end_row - start_row
                },
                'calculation_details': calculation_details
            }
            
            # 파일 타입별 추가 정보
            if file_type == 'test1' and 'unit' in item:
                price_item['unit'] = item['unit']
            elif file_type == 'est':
                price_item.update({
                    'total': item.get('total', ''),
                    'labor': item.get('labor', ''),
                    'material': item.get('material', ''),
                    'expense': item.get('expense', '')
                })
            elif file_type == 'stmate':
                price_item.update({
                    'specification': item.get('spec', ''),
                    'unit': item.get('unit', ''),
                    'total': item.get('total', ''),
                    'material': item.get('material', ''),
                    'labor': item.get('labor', ''),
                    'expense': item.get('expense', '')
                })
            elif file_type == 'construction1':
                price_item.update({
                    'material': item.get('material', ''),
                    'labor': item.get('labor', ''),
                    'expense': item.get('expense', ''),
                    'total': item.get('total', '')
                })
            
            result['price_data'].append(price_item)
        
        return result
    
    def _extract_calculation_detail(self, df: pd.DataFrame, row_idx: int, file_type: str) -> Optional[Dict]:
        """계산 상세 내용 추출"""
        if row_idx >= len(df):
            return None
        
        # 첫 번째 컬럼에서 내용 확인
        first_col = df.iloc[row_idx, 0] if len(df.columns) > 0 else None
        if not pd.notna(first_col):
            return None
        
        content = str(first_col).strip()
        
        # 빈 내용이거나 무의미한 내용 제외
        if not content or content in ['', ' ', '0']:
            return None
        
        detail = {
            'row_number': row_idx,
            'content': content
        }
        
        # 파일 타입별 추가 정보 추출
        if file_type in ['est', 'stmate', 'construction1']:
            # 재료비, 노무비, 경비 등의 정보가 있는 경우
            if len(df.columns) > 1:
                for col_idx, field_name in enumerate(['material', 'labor', 'expense', 'total'], 1):
                    if col_idx < len(df.columns):
                        cell = df.iloc[row_idx, col_idx]
                        if pd.notna(cell):
                            value = str(cell).strip()
                            if value and value != '0':
                                detail[field_name] = value
        
        return detail
    
    def _save_result(self, result: Dict, output_file: str):
        """결과 저장"""
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        # 통계 출력
        total_details = sum(len(item['calculation_details']) for item in result['price_data'])
        print(f"✅ 저장: {output_file}")
        print(f"   단가산출 항목: {result['total_price_items']}개")
        print(f"   계산 상세: {total_details}개")
    
    def run_all_price_parsing(self):
        """모든 파일의 단가산출 파싱"""
        results = []
        
        # test1.xlsx
        if os.path.exists('test1.xlsx'):
            results.append(self.parse_test1_price())
        
        # est.xlsx
        if os.path.exists('est.xlsx'):
            results.append(self.parse_est_price())
        
        # sgs.xls
        if os.path.exists('sgs.xls'):
            results.append(self.parse_sgs_price())
        
        # 건축구조내역.xlsx
        if os.path.exists('건축구조내역.xlsx'):
            results.append(self.parse_construction1_price())
        
        # stmate.xlsx (repaired)
        if os.path.exists('stmate_repaired.xlsx'):
            results.append(self.parse_stmate_price())
        
        return results

def validate_price_results():
    """단가산출 결과 검증"""
    print("\n" + "="*80)
    print("단가산출 결과 검증")
    print("="*80)
    
    files = [
        'result/test1_price.json',
        'result/est_price.json', 
        'result/sgs_price.json',
        'result/construction1_price.json',
        'result/stmate_price.json'
    ]
    
    total_success = 0
    total_files = 0
    
    for file_path in files:
        if os.path.exists(file_path):
            total_files += 1
            
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            print(f"\n{data['file']}:")
            print(f"  단가산출 항목: {data['total_price_items']}개")
            
            # 계산 상세 통계
            total_details = sum(len(item['calculation_details']) for item in data['price_data'])
            print(f"  계산 상세: {total_details}개")
            
            # 성공 판정
            success = data['total_price_items'] > 0
            if success:
                total_success += 1
                print(f"  상태: ✅ 성공")
            else:
                print(f"  상태: ❌ 실패")
    
    # 전체 성공률
    print("\n" + "="*80)
    success_rate = (total_success / total_files * 100) if total_files > 0 else 0
    
    if success_rate == 100:
        print(f"🎉 단가산출 성공률: {success_rate:.1f}% - 모든 파일 파싱 성공!")
    else:
        print(f"📊 단가산출 성공률: {success_rate:.1f}% ({total_success}/{total_files})")
    
    return success_rate

if __name__ == "__main__":
    # 단가산출 파서 실행
    parser = PriceCalculationParser()
    results = parser.run_all_price_parsing()
    
    # 검증
    success_rate = validate_price_results()
    
    if success_rate < 100:
        print("\n⚠️ 일부 파일에서 문제 발생. 추가 개선 필요.")
    else:
        print("\n✅ 모든 단가산출 파싱 성공!")